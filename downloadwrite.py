"""downloadwrite: Creates the output reports of the comparison
Can write the outputs in python files (which can be used in psse to create the new from the original.
Else the output can be in Excel/Csv format
"""
__author__ = "Sudipto Bhowmik, whit.com.au"
__license__ = "GPL"
__version__ = "1.0.0"

from collections import defaultdict
import csv
from itertools import starmap
import os.path
import sys
import sqlite3
import textwrap

# import app_settings (which sets up the rest of the paths).
try:
    import app_settings
except ImportError:
    app_dir = os.path.dirname(__file__)
    sys.path.append(app_dir)
    import app_settings

# import 3rd party packages
import openpyxl

# import Grid compare modules.
import compare
from elements import elements

# The only things that is absolutely necessary is that 'wnd' comes after
# 'tr3' and that the buses exist to be connected to.
ADD_ORDER = ['bus','genbus','load','mach','fxshunt','swsh','brn','trn',
            'tr3','wnd']

# 'wnd' element types don't have a dedicated purg function. They must get
# removed with their 'tr3'.
DEL_ORDER = ['tr3', 'trn', 'brn', 'swsh','fxshunt', 'mach', 'load', 'genbus',
                'bus']

TAB = '    '

def collect_cmds(change_rows):
    def nested_defaultdicts(n_deep, ending):
        if n_deep == 0:
            return ending
        else:
            def new_fn():
                return defaultdict(nested_defaultdicts(n_deep-1, ending))
            return new_fn

    cmds = defaultdict(nested_defaultdicts(3,dict))

    for row in change_rows:
        # MAGIC
        node, elemtype, option = row[:3]

        write_fn = elements[elemtype].writables[option].write_fn

        # MAGIC
        cmds[elemtype][node][write_fn][option] = row[3:]
    return cmds


def make_python(fid_out,origfile):
    write_py_header(fid_out)

    write_py_main(fid_out,origfile)

    fid_out.write(make_change_fn().decode('latin1'))

    write_call_main(fid_out)

def write_py_main(fid_out,origfile):
    """Write the function to be executed when the python file is executed."""

    fid_out.write(textwrap.dedent("""\
            def main():
                if not run_in_psse:
                    # Place all commands you wish to run in a Python environment,
                    # but not from within PSSE in this 'if' block.

                    # init with default number of buses.
                    psspy.psseinit()

                    # load original case
                    psspy.case("%s")

                # apply changes
                do_changes()

                # Add commands to be automatically run after changes have been
                # applied.  For instance, you may want to solve or save the new
                # case under a new name.\n\n""" % origfile)
            .decode('latin1'))


def write_call_main(fid_out):
    """Writes the final command of the file: 'main()'."""
    fid_out.write(u'\n# Execute the main function.\nmain()')

def write_py_header(fid_out):
    fid_out.write(textwrap.dedent(u"""\
        # -*- coding: utf-8 -*-
        import sys
        import os

        # Try to import psspy
        try:
            import psspy
        except ImportError, e:
            # If we couldn't load psspy out of the box, lets look for
            # pssepath, a module designed to facilitate importing psspy. The
            # quickest way of getting this to work is to place the 'pssepath'
            # python file in this directory, it will setup psspy for you.
            try:
                import pssepath
            except ImportError:
                print "\\nUnable to import psspy.\\n"
                raise e
            else:
                pssepath.add_pssepath()
                import psspy

        # Determine if this script is being run from PSSE or plain Python.
        executable = os.path.basename(sys.executable).lower()
        run_in_psse = True
        if executable in ['python.exe', 'pythonw.exe']:
            run_in_psse = False

        psspy.throwPsseExceptions = True
        \n
        """).decode('latin1'))


def make_change_fn():
    # remove elements

    lines = []

    fn_sig = "def do_changes():"

    removed = compare.unique_elems_by_type('removed')
    lines.extend(removed_lines(removed, indent=TAB))

    # new elements
    new = compare.changes_by_type('new')
    lines.extend(new_lines(new, indent=TAB))

    # changed elements
    changes = compare.changes_by_bus()
    lines.extend(changed_lines(changes, indent=TAB))

    lines.insert(0,fn_sig)

    return '\n'.join(lines)

def removed_lines(removed_elems_by_type, indent=''):

    lines = []
    lines.append(indent + '# ====== Removed Elements =====\n')
    nodejoin = ',\n%s%s' % (indent, TAB)

    for elemtype in DEL_ORDER:
        lines.append(indent + '# === Removed %s' % elemtype)
        nodes = removed_elems_by_type[elemtype]
        e = elements[elemtype]
        cmd_strs = []
        if elemtype == 'bus':
            # make an sid of all the buses that need to be removed.
            if nodes:
                sid_str = indent + 'psspy.bsys(11, numbus=%s, buses=[%s])'
                cmd_strs.append(sid_str % (len(nodes), nodejoin.join(nodes)))
                cmd_strs.append(indent + 'psspy.extr(sid=11)')

        elif elemtype == 'genbus':
            # can't use purgplnt if the machines have already been removed.
            if nodes:
                sid_str = indent + 'psspy.bsys(11, numbus=%s, buses=[%s])'
                cmd_strs.append(sid_str % (len(nodes), nodejoin.join(nodes)))
                # opt = 4 is plants
                cmd_strs.append(indent + 'psspy.purg(sid=11, opt=4)')

        else:
            # every other element.
            for node in nodes:
                if e.del_fn:
                    cmd_strs.append(
                            '\n'.join(e.del_fn.write_psspy_cmd(node,
                                initial_indent=TAB,
                                subsequent_indent=TAB*2)
                            ))

        lines.extend(cmd_strs)
        lines.append('\n')

    return lines

def new_lines(added, indent=''):

    lines = []

    lines.append(indent + '# ====== Added Elements =====\n')

    for elemtype in ADD_ORDER:
        add_cmds = collect_cmds(added[elemtype])
        lines.append(indent + '# === Added %s' % elemtype)
        nodes = add_cmds[elemtype]
        e = elements[elemtype]
        cmd_strs = []
        for node, write_fns in nodes.items():
            for write_fn, options in write_fns.items():
                fn = e.write_fns[write_fn]
                cmd_strs.extend(fn.write_psspy_cmd(node, options,
                    initial_indent=TAB, subsequent_indent=TAB*2))

        lines.extend(cmd_strs)
        lines.append('\n')

    return lines

def changed_lines(changes, indent=''):

    lines = []
    lines.append(indent + '# ====== Changed Elements =====\n')

    for bus, change_rows in changes.items():
        lines.append(indent + "# Bus %d --------" % (bus,))
        cmd_strs = []
        collect_changes = collect_cmds(change_rows)
        for elemtype in ADD_ORDER:
            nodes = collect_changes[elemtype]
            e = elements[elemtype]
            for node, write_fns in nodes.items():
                for write_fn, options in write_fns.items():
                    fn = e.write_fns[write_fn]
                    cmd_strs.extend(fn.write_psspy_cmd(node, options,
                        initial_indent=TAB, subsequent_indent=TAB*2))

        lines.extend(cmd_strs)
        lines.append('\n')

    return lines


def make_csv(fid):
    """Write out a csv file of all the changes."""
    writer = csv.writer(fid)

    con = sqlite3.connect(app_settings.COMPARE_DB)
    con.text_factory = str

    # --- Write removed data
    del_sql = """SELECT node, type, option, value
                    FROM removed"""

    # change option to a more readable name.
    rows = [(node, elem_type,
        elements[elem_type].writables[option].read_param,
        value.strip(),'')
        for node, elem_type, option, value in con.execute(del_sql)]

    writer.writerows(rows)

    # --- Write new data
    new_sql = """SELECT node, type, option, value
                    FROM new"""
    # change option to a more readable name.
    rows = [(node, elem_type,
        elements[elem_type].writables[option].read_param,
        '', value.strip())
        for node, elem_type, option, value in con.execute(new_sql)]

    writer.writerows(rows)

    # --- Write changed data
    comp_sql = """SELECT node, type, option, before, after
                    FROM compare"""
    # change option to a more readable name.
    rows = [(node, elem_type,
        elements[elem_type].writables[option].read_param,
        before.strip(), after.strip())
        for node, elem_type, option, before, after in con.execute(comp_sql)]

    writer.writerows(rows)

# ============== Excel Writer ============

def excel_write_rows(sheet, data_rows):
    """Writes the data stored in 'data_rows' to an excel spreadsheet 'sheet'.

    Returns the number of lines written.
    """

    # If data_rows is empty, the loop won't run and 'i' is undefined. This
    # makes the fn return 0 lines in that case.
    for row in data_rows:
        sheet.append(row)

def write_excel_sheet_header(sheet, e):
    """Returns a tuple with the names to be used as the column headers."""
    a_write_fn = e.write_fns.values()[0]
    headers = a_write_fn.primaries.keys()

    headers += ('option', 'case A', 'case B')

    sheet.append(headers)
    # letters = ['a','b','c','d','e','f','g','h','i','j','k','m','n','o']
    # for cell, header_val in zip(
    #         sheet.range('A1:%s1' % letters[len(headers)-1])[0],
    #         headers):
    #     cell.value = header_val
    #     cell.style.font.bold = True

def construct_data_format_fn(e, action):
    """Returns a function for converting a row from the database into a row to
    be used for excel.

    The conversion function is required to convert the node id into its
    component values and convert floats and ints which are strings to the
    appropriate type.

    'action' is a string which determines the type of row to be returned. This
    is because the row has to have a before and after value and this is
    different for the three cases.
    'action':   'changed' - changed style row
                'added'   - added style row
                'removed' - removed style row.
    """
    a_write_fn = e.write_fns.values()[0]

    get_vals_from_node = a_write_fn.get_keyword_vals_from_nodeid

    type_fn_dict = {'int':int, 'real':float,
                     'char': lambda x: x.decode('latin1').strip()}
    type_conv = []
    for primary in a_write_fn.primaries.values():
        type_conv.append(type_fn_dict[primary.data_type])

    if action == 'changed':
        def row_db_to_excel_format(node, option, before, after):
            """Takes a row returned from the sql query and prepares the data
            for entry into excel.

            See the construct_data_format_fn docstring for more info."""
            node_vals = get_vals_from_node(node)
            node_vals = tuple([f(x) for f, x in zip(type_conv, node_vals)])

            conv_fn = type_fn_dict[e.writables[option].data_type]

            opt_and_vals = (e.writables[option].read_param,
                conv_fn(before),
                conv_fn(after))

            return node_vals + opt_and_vals

    elif action == 'added':
        def row_db_to_excel_format(node, option, value):
            """Takes a row returned from the sql query and prepares the data
            for entry into excel.

            See the construct_data_format_fn docstring for more info."""
            node_vals = get_vals_from_node(node)
            node_vals = tuple([f(x) for f, x in zip(type_conv, node_vals)])
            try:
                opt_and_vals = (e.writables[option].read_param, '',
                    type_fn_dict[e.writables[option].data_type](value))
            except TypeError:
                opt_and_vals = (e.writables[option].read_param, '','')

            return node_vals + opt_and_vals

    elif action == 'removed':
        def row_db_to_excel_format(node, option, value):
            """Takes a row returned from the sql query and prepares the data
            for entry into excel.

            See the construct_data_format_fn docstring for more info."""
            node_vals = get_vals_from_node(node)
            node_vals = tuple([f(x) for f, x in zip(type_conv, node_vals)])
            try:
                opt_and_vals = (e.writables[option].read_param,
                    type_fn_dict[e.writables[option].data_type](value), '')
            except TypeError:
                opt_and_vals = (e.writables[option].read_param,'','')

            return node_vals + opt_and_vals


    return row_db_to_excel_format

def make_excel(filename):
    """Write an excel spreadsheet to fid."""
    wb = openpyxl.Workbook()

    con = sqlite3.connect(app_settings.COMPARE_DB)
    con.text_factory = str

    sheetorder = ['bus','genbus','mach','load', 'fxshunt', 'swsh', 'brn',
            'trn', 'tr3', 'wnd']

    elem_to_sheet_name = {'bus':'Bus',
            'genbus':'Plant',
            'mach':'Machine',
            'load':'Load',
            'fxshunt':'Fixed Shunt',
            'swsh':'Switch Shunt',
            'brn':'Branch',
            'trn':'2 Trn',
            'tr3':'3 Trn',
            'wnd':'3 Trn Windings'}

    type_conv = {'int':int, 'real':float, 'char': lambda x: x.decode('latin1')}
    sheets = {}
    for elem_name in sheetorder:
        sheet_name = elem_to_sheet_name[elem_name]
        e = elements[elem_name]

        del_sql = """SELECT node, option, value
                    FROM removed
                    WHERE type='%s'""" % elem_name
        # --- Count removed data
        del_count, = con.execute("""SELECT count(*) FROM removed
                                    WHERE type='%s'""" % elem_name).fetchone()

        new_sql = """SELECT node, option, value
                    FROM new
                    WHERE type='%s'""" % elem_name
        # --- Count added data
        new_count, = con.execute("""SELECT count(*) FROM new
                                    WHERE type='%s'""" % elem_name).fetchone()

        comp_sql = """SELECT node, option, before, after
                        FROM compare
                        WHERE type='%s'""" % elem_name
        # --- Count changed data
        comp_count, = con.execute("""SELECT count(*) FROM compare
                                    WHERE type='%s'""" % elem_name).fetchone()

        if del_count or new_count or comp_count:
            # There are changes so lets make a worksheet.
            e_sheet = wb.create_sheet(title=sheet_name)
            sheets[sheet_name] = e_sheet

            # write the headers on the 0th line.
            write_excel_sheet_header(e_sheet, e)

            # --- Write removed data
            del_rows = starmap(construct_data_format_fn(e,'removed'),
                    con.execute(del_sql))
            excel_write_rows(e_sheet, del_rows)

            # --- Write added data
            new_rows = starmap(construct_data_format_fn(e,'added'),
                    con.execute(new_sql))
            excel_write_rows(e_sheet, new_rows)

            # --- Write changed data
            comp_rows = starmap(construct_data_format_fn(e,'changed'),
                    con.execute(comp_sql))
            excel_write_rows(e_sheet, comp_rows)

    if not sheets:
        # there are no changes between the 2 files.
        sheet = wb.create_sheet(title='Info')
        sheet.append(['No changes found between the two files '
                'for the elements that were checked.'])

        sheet.append(['The following network types were checked:'])
        for network_type in sheetorder:
            sheet.append([elem_to_sheet_name[network_type]])

    # Remove the default sheet created with the workbook.
    wb.remove_sheet(wb.worksheets[0])
    wb.save(filename)
    return True
